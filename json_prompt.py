import os
import json
import ast
import itertools
from collections import ChainMap
from execution import PromptExecutor
from . import register_node
from PIL import Image, ImageOps, ImageSequence, ImageFile
from PIL.PngImagePlugin import PngInfo
import folder_paths
import latent_preview
import node_helpers
import numpy as np
import safetensors.torch
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment


from folder_paths import *
folder_names_and_paths["image"] = ([os.path.join(models_dir, "image")], [".json"])
@register_node
class ImageInputNode:
    def __init__(self):
        self.images = []

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": { 
                "json_file": (folder_paths.get_filename_list("image"), ),
            }
        }

    RETURN_TYPES = ("SEQUENCE",)
    RETURN_NAMES = ("sequence", )
    FUNCTION = "get_next_image"
    CATEGORY = "test pipeline"

    def load_images_from_file(self, file_path):
        full_path = file_path
        with open(full_path, 'r') as file:
            data = json.load(file)
            self.images = [image['path'] for image in data.get('images', [])]

        if not self.images:
            raise ValueError(f"No prompts found in the JSON file: {full_path}")

    def get_next_image(self, json_file):
        if not self.images:
            self.load_images_from_file(json_file)
        
        images_as_string = str(self.images)
        
        return (ast.literal_eval(images_as_string), )

from folder_paths import *
folder_names_and_paths["prompt"] = ([os.path.join(models_dir, "prompt")], [".json"])

@register_node
class JSONPromptNode:
    
    def __init__(self):
        self.prompts = []

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": { 
                "json_file": (folder_paths.get_filename_list("prompt"), ),
            }
        }

    RETURN_TYPES = ("SEQUENCE",)
    RETURN_NAMES = ("sequence", )
    FUNCTION = "get_next_prompt"
    CATEGORY = "test pipeline"

    def load_prompts_from_file(self, file_path):
        full_path = folder_paths.get_full_path("prompt", file_path)
        with open(full_path, 'r') as file:
            data = json.load(file)
            self.prompts = [prompt['text'] for prompt in data.get('prompts', [])]

        if not self.prompts:
            raise ValueError(f"No prompts found in the JSON file: {full_path}")

    def get_next_prompt(self, json_file):
        if not self.prompts:
            self.load_prompts_from_file(json_file)
        
        prompts_as_string = str(self.prompts)
        
        return (ast.literal_eval(prompts_as_string), )


@register_node
class MakeJob:
    """Turns a sequence into a job with one attribute."""

    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "sequence": ("SEQUENCE", ),
                "name": ("STRING", {"default": ''}),
            },
        }

    RETURN_TYPES = ("JOB", "INT")
    RETURN_NAMES = ("job", "count")
    FUNCTION = "go"
    CATEGORY = "test pipeline"

    def merge_dicts(self, *dicts):
        return dict(itertools.chain.from_iterable(d.items() for d in dicts))

    def go(self, sequence, name):
        result = [{name: value} for value in sequence]
        return (result, len(result))
    
@register_node
class ExtractImages:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "attributes": ("ATTRIBUTES",)
            },
        }
    
    RETURN_TYPES = ("IMAGE", "MASK")
    RETURN_NAMES = ("image", "mask")
    FUNCTION = "go"
    CATEGORY = "test pipeline"
    
    def go(self,attributes):
        image_path = next(iter(attributes.values()))
        img = node_helpers.pillow(Image.open, image_path)
        
        output_images = []
        output_masks = []
        w, h = None, None

        excluded_formats = ['MPO']
        
        for i in ImageSequence.Iterator(img):
            i = node_helpers.pillow(ImageOps.exif_transpose, i)

            if i.mode == 'I':
                i = i.point(lambda i: i * (1 / 255))
            image = i.convert("RGB")

            if len(output_images) == 0:
                w = image.size[0]
                h = image.size[1]
            
            if image.size[0] != w or image.size[1] != h:
                continue
            
            image = np.array(image).astype(np.float32) / 255.0
            image = torch.from_numpy(image)[None,]
            if 'A' in i.getbands():
                mask = np.array(i.getchannel('A')).astype(np.float32) / 255.0
                mask = 1. - torch.from_numpy(mask)
            else:
                mask = torch.zeros((64,64), dtype=torch.float32, device="cpu")
            output_images.append(image)
            output_masks.append(mask.unsqueeze(0))

        if len(output_images) > 1 and img.format not in excluded_formats:
            output_image = torch.cat(output_images, dim=0)
            output_mask = torch.cat(output_masks, dim=0)
        else:
            output_image = output_images[0]
            output_mask = output_masks[0]

        return (output_image, output_mask)
    
@register_node
class FormatAttributes:
    """Applies attributes to a format string."""
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "attributes": ("ATTRIBUTES",),
                "format": ("STRING", {'default': '', 'multiline': True, "dynamicPrompts": False})
            },
        }

    RETURN_TYPES = ("STRING", )
    RETURN_NAMES = ("string", )
    FUNCTION = "go"
    CATEGORY = "test pipeline"

    def go(self, attributes, format):
        return (format.format(**attributes), )

@register_node
class JobIterator:
    """Magic node that runs the workflow multiple times until all steps are done."""
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "job": ("JOB",),
                "start_step": ("INT", {"default": 0}),
            },
        }

    RETURN_TYPES = ("ATTRIBUTES", "INT", "INT")
    RETURN_NAMES = ("attributes", "count", "step")
    FUNCTION = "go"
    CATEGORY = "test pipeline"

    def go(self, job, start_step):
        print(f'JobIterator: {start_step + 1} / {len(job)}')
        return (job[start_step], len(job), start_step)
    
@register_node
class CustomSaveImage:
    def __init__(self):
        self.type = "output"
    
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
              "images": ("IMAGE",),
              "dir_name": ("STRING", {"default": ''}),
              "prompt": ("STRING", {"default": ''}),
            },
        }
    
    RETURN_TYPES = ()
    RETURN_NAMES = ()
    FUNCTION = "go"
    CATEGORY = "test pipeline"
    OUTPUT_NODE = True

    def go(self, images, dir_name, prompt):
        results = list()
        output_dir_path = os.path.join(folder_paths.base_path, 'output', dir_name)

        os.makedirs(output_dir_path, exist_ok=True)

        excel_path = os.path.join(output_dir_path, f"{dir_name}.xlsx")

        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = dir_name
            ws.column_dimensions['A'].width = 150
            ws.column_dimensions['B'].width = 100
            wb.save(excel_path)
        else:
            wb = load_workbook(excel_path)
            ws = wb.active

        max_filename_length = 50 
        truncated_prompt = prompt[:max_filename_length - len("_.png")]

        for (batch_number, image) in enumerate(images):
            i = 255. * image.cpu().numpy()
            img = Image.fromarray(np.clip(i, 0, 255).astype(np.uint8))
            file = f"{truncated_prompt}_.png"
            file_path = os.path.join(output_dir_path, file)
            img.save(file_path)
            results.append({
                "filename": file,
                "subfolder": dir_name,
                "type": self.type
            })

            next_row = ws.max_row + 1
            ws.row_dimensions[next_row].height = 800
            ws[f'B{next_row}'] = prompt
            ws[f'B{next_row}'].alignment = Alignment(horizontal='left', vertical='top')

            excel_img = OpenpyxlImage(file_path)
            ws.add_image(excel_img, f'A{next_row}')

        wb.save(excel_path)

        return { "ui": { "images": results } }
    
orig_execute = PromptExecutor.execute


def execute(self, prompt, prompt_id, extra_data={}, execute_outputs=[]):
    print("Prompt executor has been patched by Job Iterator!")
    orig_execute(self, prompt, prompt_id, extra_data, execute_outputs)

    steps = None
    job_iterator = None
    for k, v in prompt.items():
        try:
            if v['class_type'] == 'JobIterator':
                steps = self.outputs[k][1][0]
                job_iterator = k
                break
        except KeyError:
            continue
    else:
        return

    while prompt[job_iterator]['inputs']['start_step'] < (steps - 1):
        prompt[job_iterator]['inputs']['start_step'] += 1
        orig_execute(self, prompt, prompt_id, extra_data, execute_outputs)


PromptExecutor.execute = execute
